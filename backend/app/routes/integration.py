"""Integration ingestion routes."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from http import HTTPStatus
from math import ceil
from typing import Any, Dict, List

from flask import Blueprint, jsonify, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import func, or_, select
from sqlalchemy.exc import SQLAlchemyError

from ..database import session_scope
from ..services.configuration import get_snapshot
from ..services.ingestion import (
    build_envelope,
    log_envelope,
    persist_integration,
    update_integration_record,
)
from ..models import IntegrationRecord

integration_bp = Blueprint("integration", __name__, url_prefix="/api/integration")


_SORTABLE_FIELDS = {
    "id": IntegrationRecord.id,
    "matricula": IntegrationRecord.matricula,
    "nome": IntegrationRecord.nome,
    "setor": IntegrationRecord.setor,
    "cargo": IntegrationRecord.cargo,
    "turno": IntegrationRecord.turno,
    "integracao": IntegrationRecord.integracao,
    "supervisor": IntegrationRecord.supervisor,
    "data": IntegrationRecord.data,
    "submitted_at": IntegrationRecord.submitted_at,
}


def _build_integration_filters(search_term: str) -> List[Any]:
    filters: List[Any] = []
    if search_term:
        like_pattern = f"%{search_term}%"
        filters.append(
            or_(
                IntegrationRecord.matricula.ilike(like_pattern),
                IntegrationRecord.nome.ilike(like_pattern),
                IntegrationRecord.setor.ilike(like_pattern),
                IntegrationRecord.cargo.ilike(like_pattern),
                IntegrationRecord.turno.ilike(like_pattern),
                IntegrationRecord.integracao.ilike(like_pattern),
                IntegrationRecord.supervisor.ilike(like_pattern),
            )
        )
    return filters


def _format_date(value) -> str:
    if not value:
        return ""
    return value.strftime("%d/%m/%Y")


def _format_datetime(value) -> str:
    if not value:
        return ""
    return value.strftime("%d/%m/%Y %H:%M")


def _auto_fit_columns(sheet) -> None:
    for index, column in enumerate(sheet.columns, start=1):
        max_length = 0
        for cell in column:
            value = cell.value
            length = len(str(value)) if value is not None else 0
            max_length = max(max_length, length)
        adjusted = min(max_length + 4, 48)
        sheet.column_dimensions[get_column_letter(index)].width = max(adjusted, 12)


def _build_integration_workbook(records: List[IntegrationRecord]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Integrações"

    headers = [
        "ID",
        "Matrícula",
        "Colaborador",
        "Setor",
        "Cargo",
        "Turno",
        "Integração",
        "Supervisor",
        "Data integração",
        "Observação",
        "Registrado em",
    ]
    sheet.append(headers)

    for record in records:
        sheet.append(
            [
                record.id,
                record.matricula or "",
                record.nome,
                record.setor,
                record.cargo,
                record.turno,
                record.integracao,
                record.supervisor,
                _format_date(record.data),
                record.observacao or "",
                _format_datetime(record.submitted_at),
            ]
        )

    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="2563EB")
    header_alignment = Alignment(vertical="center", horizontal="left")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}1"
    _auto_fit_columns(sheet)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def _send_workbook(stream: BytesIO, filename: str):
    """Return a Flask file response compatible with older Flask releases."""
    mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    try:
        return send_file(stream, mimetype=mimetype, as_attachment=True, download_name=filename)
    except TypeError:
        stream.seek(0)
        return send_file(stream, mimetype=mimetype, as_attachment=True, attachment_filename=filename)


@integration_bp.get("")
def describe_submission() -> Any:
    """Expose metadata about the integration submission endpoint."""
    with session_scope() as session:
        options = get_snapshot(session).get("integration", {})

    payload = {
        "service": "integration-backend",
        "description": "Recebe cargas do formulário de integração de colaboradores.",
        "accepted_methods": ["POST"],
        "endpoint": "/api/integration",
        "content_type": "application/json",
        "payload_expectation": "Objeto JSON com dados de colaboradores e metadados opcionais.",
        "options": options,
    }

    return jsonify(payload), HTTPStatus.OK


@integration_bp.get("/export")
def export_records() -> Any:
    search_param = request.args.get("search", default="").strip()
    sort_by_param = request.args.get("sort_by", default="submitted_at")
    sort_order_param = request.args.get("sort_order", default="desc")

    with session_scope() as session:
        filters = _build_integration_filters(search_param)
        data_stmt = select(IntegrationRecord)

        if filters:
            data_stmt = data_stmt.where(*filters)

        sort_column = _SORTABLE_FIELDS.get(sort_by_param, IntegrationRecord.submitted_at)
        sort_order = sort_order_param.lower()
        order_clause = sort_column.asc() if sort_order == "asc" else sort_column.desc()
        data_stmt = data_stmt.order_by(order_clause)

        records = session.execute(data_stmt).scalars().all()

    stream = _build_integration_workbook(records)
    filename = f"integracoes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return _send_workbook(stream, filename)


@integration_bp.post("")
def accept_submission() -> Any:
    """Receive integration data from the frontend.

    The payload is stored server-side only for observability at this phase.
    """
    payload = _extract_payload()
    if payload is None:
        return jsonify({"error": "Não foi possível interpretar os dados enviados."}), HTTPStatus.BAD_REQUEST

    try:
        with session_scope() as session:
            record = persist_integration(session, payload)
            options = get_snapshot(session).get("integration", {})
            envelope = build_envelope(_record_payload(record), origin="frontend")
            log_envelope(envelope, record_id=record.id)

        return (
            jsonify(
                {
                    "status": "accepted",
                    "record_id": record.id,
                    "submission_id": envelope.submission_id,
                    "received_at": envelope.received_at.isoformat(),
                    "options": options,
                }
            ),
            HTTPStatus.ACCEPTED,
        )
    except ValueError as error:
        return jsonify({"error": str(error)}), HTTPStatus.BAD_REQUEST
    except SQLAlchemyError:
        return jsonify({"error": "Erro ao salvar a integração."}), HTTPStatus.INTERNAL_SERVER_ERROR


@integration_bp.put("/records/<int:record_id>")
def update_record(record_id: int) -> Any:
    payload = _extract_payload()
    if payload is None:
        return jsonify({"error": "Não foi possível interpretar os dados enviados."}), HTTPStatus.BAD_REQUEST

    try:
        with session_scope() as session:
            record = session.get(IntegrationRecord, record_id)
            if record is None:
                return jsonify({"error": "Registro de integração não encontrado."}), HTTPStatus.NOT_FOUND

            updated_record = update_integration_record(session, record, payload)
            options = get_snapshot(session).get("integration", {})
            envelope = build_envelope(_record_payload(updated_record), origin="frontend")
            log_envelope(envelope, record_id=updated_record.id)

        return (
            jsonify(
                {
                    "status": "updated",
                    "record_id": record_id,
                    "options": options,
                }
            ),
            HTTPStatus.OK,
        )
    except ValueError as error:
        return jsonify({"error": str(error)}), HTTPStatus.BAD_REQUEST
    except SQLAlchemyError:
        return jsonify({"error": "Erro ao atualizar a integração."}), HTTPStatus.INTERNAL_SERVER_ERROR


@integration_bp.delete("/records/<int:record_id>")
def delete_record(record_id: int) -> Any:
    try:
        with session_scope() as session:
            record = session.get(IntegrationRecord, record_id)
            if record is None:
                return jsonify({"error": "Registro de integração não encontrado."}), HTTPStatus.NOT_FOUND

            session.delete(record)

        return jsonify({"status": "deleted", "record_id": record_id}), HTTPStatus.OK
    except SQLAlchemyError:
        return jsonify({"error": "Erro ao remover a integração."}), HTTPStatus.INTERNAL_SERVER_ERROR


@integration_bp.get("/records")
def list_records() -> Any:
    page_param = request.args.get("page", default="1")
    size_param = request.args.get("page_size", default="10")
    sort_by_param = request.args.get("sort_by", default="submitted_at")
    sort_order_param = request.args.get("sort_order", default="desc")
    search_param = request.args.get("search", default="").strip()

    try:
        page = max(int(page_param), 1)
    except (TypeError, ValueError):
        page = 1

    try:
        page_size = max(min(int(size_param), 50), 1)
    except (TypeError, ValueError):
        page_size = 10

    with session_scope() as session:
        filters = _build_integration_filters(search_param)

        count_stmt = select(func.count()).select_from(IntegrationRecord)
        data_stmt = select(IntegrationRecord)

        if filters:
            count_stmt = count_stmt.where(*filters)
            data_stmt = data_stmt.where(*filters)

        total_items = session.execute(count_stmt).scalar_one()

        total_pages = ceil(total_items / page_size) if total_items else 0
        if total_pages:
            page = min(page, total_pages)
        else:
            page = 1

        offset = (page - 1) * page_size

        sort_column = _SORTABLE_FIELDS.get(sort_by_param, IntegrationRecord.submitted_at)
        sort_order = sort_order_param.lower()
        order_clause = sort_column.asc() if sort_order == "asc" else sort_column.desc()

        data_stmt = data_stmt.order_by(order_clause).offset(offset).limit(page_size)
        records = session.execute(data_stmt).scalars().all()

    payload = {
        "items": [_serialize_integration_record(record) for record in records],
        "pagination": {
            "page": page,
            "page_size": page_size,
            "total_items": total_items,
            "total_pages": total_pages or 1,
        },
    }

    return jsonify(payload), HTTPStatus.OK


def _serialize_integration_record(record: IntegrationRecord) -> Dict[str, Any]:
    return {
        "id": record.id,
        "matricula": record.matricula,
        "nome": record.nome,
        "setor": record.setor,
        "cargo": record.cargo,
        "turno": record.turno,
        "integracao": record.integracao,
        "supervisor": record.supervisor,
        "data": record.data.isoformat() if record.data else None,
        "observacao": record.observacao,
        "submitted_at": record.submitted_at.isoformat(),
    }


def _extract_payload() -> Dict[str, Any] | None:
    if request.is_json:
        data = request.get_json(silent=True)
        if isinstance(data, dict):
            return data
    form_data = request.form.to_dict(flat=True)
    return form_data or None


def _record_payload(record: IntegrationRecord) -> Dict[str, Any]:
    return {
        "submission_id": record.id,
        "matricula": record.matricula,
        "nome": record.nome,
        "setor": record.setor,
        "integracao": record.integracao,
        "supervisor": record.supervisor,
        "turno": record.turno,
        "cargo": record.cargo,
        "data": record.data.isoformat() if record.data else None,
        "observacao": record.observacao,
        "submitted_at": record.submitted_at.isoformat(),
    }
