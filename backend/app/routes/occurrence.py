"""Occurrence ingestion routes."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from math import ceil
from http import HTTPStatus
from typing import Any, Dict, List

from flask import Blueprint, jsonify, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import func, or_, select
from sqlalchemy.exc import SQLAlchemyError

from ..database import session_scope
from ..models import OccurrenceRecord
from ..services.configuration import get_snapshot
from ..services.ingestion import persist_occurrence, update_occurrence_record

occurrence_bp = Blueprint("occurrence", __name__, url_prefix="/api/occurrence")


_SORTABLE_FIELDS = {
    "id": OccurrenceRecord.id,
    "matricula": OccurrenceRecord.matricula,
    "nome": OccurrenceRecord.nome,
    "setor": OccurrenceRecord.setor,
    "cargo": OccurrenceRecord.cargo,
    "turno": OccurrenceRecord.turno,
    "motivo": OccurrenceRecord.motivo,
    "grau": OccurrenceRecord.grau,
    "volumes": OccurrenceRecord.volumes,
    "supervisor": OccurrenceRecord.supervisor,
    "created_at": OccurrenceRecord.created_at,
}


def _build_occurrence_filters(search_term: str) -> List[Any]:
    filters: List[Any] = []
    if search_term:
        like_pattern = f"%{search_term}%"
        filters.append(
            or_(
                OccurrenceRecord.matricula.ilike(like_pattern),
                OccurrenceRecord.nome.ilike(like_pattern),
                OccurrenceRecord.setor.ilike(like_pattern),
                OccurrenceRecord.cargo.ilike(like_pattern),
                OccurrenceRecord.turno.ilike(like_pattern),
                OccurrenceRecord.motivo.ilike(like_pattern),
                OccurrenceRecord.supervisor.ilike(like_pattern),
                OccurrenceRecord.observacao.ilike(like_pattern),
            )
        )
    return filters


def _format_datetime(value) -> str:
    if not value:
        return ""
    return value.strftime("%d/%m/%Y %H:%M")


def _auto_fit_columns(sheet) -> None:
    for index, column in enumerate(sheet.columns, start=1):
        max_length = 0
        for cell in column:
            value = cell.value
            length = len(str(value)) if value is not None else 0
            max_length = max(max_length, length)
        adjusted = min(max_length + 4, 48)
        sheet.column_dimensions[get_column_letter(index)].width = max(adjusted, 12)


def _build_occurrence_workbook(records: List[OccurrenceRecord]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Ocorrências"

    headers = [
        "ID",
        "Matrícula",
        "Colaborador",
        "Setor",
        "Cargo",
        "Turno",
        "Supervisor",
        "Motivo",
        "Grau",
        "Grau (valor)",
        "Volumes",
        "Observação",
        "Registrado em",
    ]
    sheet.append(headers)

    for record in records:
        sheet.append(
            [
                record.id,
                record.matricula or "",
                record.nome,
                record.setor,
                record.cargo,
                record.turno,
                record.supervisor,
                record.motivo or "",
                record.grau_label or "",
                record.grau if record.grau is not None else "",
                record.volumes if record.volumes is not None else "",
                record.observacao or "",
                _format_datetime(record.created_at),
            ]
        )

    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="DC2626")
    header_alignment = Alignment(vertical="center", horizontal="left")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}1"
    _auto_fit_columns(sheet)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def _send_workbook(stream: BytesIO, filename: str):
    """Return a Flask file response compatible with older Flask releases."""
    mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    try:
        return send_file(stream, mimetype=mimetype, as_attachment=True, download_name=filename)
    except TypeError:
        stream.seek(0)
        return send_file(stream, mimetype=mimetype, as_attachment=True, attachment_filename=filename)


@occurrence_bp.post("")
def register_occurrence() -> Any:
    payload = _extract_payload()
    if payload is None:
        return jsonify({"error": "Não foi possível interpretar os dados enviados."}), HTTPStatus.BAD_REQUEST

    try:
        with session_scope() as session:
            record = persist_occurrence(session, payload)
            options = get_snapshot(session).get("occurrence", {})
        summary = _record_summary(record)
        print("\n=== Ocorrência registrada ===", flush=True)
        print(summary, flush=True)
        print("=== Fim da ocorrência ===\n", flush=True)
        return (
            jsonify(
                {
                    "status": "accepted",
                    "record_id": record.id,
                    "options": options,
                }
            ),
            HTTPStatus.ACCEPTED,
        )
    except ValueError as error:
        return jsonify({"error": str(error)}), HTTPStatus.BAD_REQUEST
    except SQLAlchemyError:
        return jsonify({"error": "Erro ao salvar a ocorrência."}), HTTPStatus.INTERNAL_SERVER_ERROR


@occurrence_bp.put("/records/<int:record_id>")
def update_occurrence(record_id: int) -> Any:
    payload = _extract_payload()
    if payload is None:
        return jsonify({"error": "Não foi possível interpretar os dados enviados."}), HTTPStatus.BAD_REQUEST

    try:
        with session_scope() as session:
            record = session.get(OccurrenceRecord, record_id)
            if record is None:
                return jsonify({"error": "Registro de ocorrência não encontrado."}), HTTPStatus.NOT_FOUND

            updated_record = update_occurrence_record(session, record, payload)
            options = get_snapshot(session).get("occurrence", {})

        return (
            jsonify(
                {
                    "status": "updated",
                    "record_id": record_id,
                    "options": options,
                    "grau_label": updated_record.grau_label,
                }
            ),
            HTTPStatus.OK,
        )
    except ValueError as error:
        return jsonify({"error": str(error)}), HTTPStatus.BAD_REQUEST
    except SQLAlchemyError:
        return jsonify({"error": "Erro ao atualizar a ocorrência."}), HTTPStatus.INTERNAL_SERVER_ERROR


@occurrence_bp.delete("/records/<int:record_id>")
def delete_occurrence(record_id: int) -> Any:
    try:
        with session_scope() as session:
            record = session.get(OccurrenceRecord, record_id)
            if record is None:
                return jsonify({"error": "Registro de ocorrência não encontrado."}), HTTPStatus.NOT_FOUND

            session.delete(record)

        return jsonify({"status": "deleted", "record_id": record_id}), HTTPStatus.OK
    except SQLAlchemyError:
        return jsonify({"error": "Erro ao remover a ocorrência."}), HTTPStatus.INTERNAL_SERVER_ERROR


@occurrence_bp.get("/records")
def list_occurrences() -> Any:
    page_param = request.args.get("page", default="1")
    size_param = request.args.get("page_size", default="10")
    sort_by_param = request.args.get("sort_by", default="created_at")
    sort_order_param = request.args.get("sort_order", default="desc")
    search_param = request.args.get("search", default="").strip()

    try:
        page = max(int(page_param), 1)
    except (TypeError, ValueError):
        page = 1

    try:
        page_size = max(min(int(size_param), 50), 1)
    except (TypeError, ValueError):
        page_size = 10

    with session_scope() as session:
        filters = _build_occurrence_filters(search_param)

        count_stmt = select(func.count()).select_from(OccurrenceRecord)
        data_stmt = select(OccurrenceRecord)

        if filters:
            count_stmt = count_stmt.where(*filters)
            data_stmt = data_stmt.where(*filters)

        total_items = session.execute(count_stmt).scalar_one()

        total_pages = ceil(total_items / page_size) if total_items else 0
        if total_pages:
            page = min(page, total_pages)
        else:
            page = 1

        offset = (page - 1) * page_size

        sort_column = _SORTABLE_FIELDS.get(sort_by_param, OccurrenceRecord.created_at)
        sort_order = sort_order_param.lower()
        order_clause = sort_column.asc() if sort_order == "asc" else sort_column.desc()

        data_stmt = data_stmt.order_by(order_clause).offset(offset).limit(page_size)
        records = session.execute(data_stmt).scalars().all()

    payload = {
        "items": [_serialize_occurrence_record(record) for record in records],
        "pagination": {
            "page": page,
            "page_size": page_size,
            "total_items": total_items,
            "total_pages": total_pages or 1,
        },
    }

    return jsonify(payload), HTTPStatus.OK


@occurrence_bp.get("/export")
def export_occurrences() -> Any:
    search_param = request.args.get("search", default="").strip()
    sort_by_param = request.args.get("sort_by", default="created_at")
    sort_order_param = request.args.get("sort_order", default="desc")

    with session_scope() as session:
        filters = _build_occurrence_filters(search_param)
        data_stmt = select(OccurrenceRecord)

        if filters:
            data_stmt = data_stmt.where(*filters)

        sort_column = _SORTABLE_FIELDS.get(sort_by_param, OccurrenceRecord.created_at)
        sort_order = sort_order_param.lower()
        order_clause = sort_column.asc() if sort_order == "asc" else sort_column.desc()
        data_stmt = data_stmt.order_by(order_clause)

        records = session.execute(data_stmt).scalars().all()

    stream = _build_occurrence_workbook(records)
    filename = f"ocorrencias_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return _send_workbook(stream, filename)


def _extract_payload() -> Dict[str, Any] | None:
    if request.is_json:
        data = request.get_json(silent=True)
        if isinstance(data, dict):
            return data
    form_data = request.form.to_dict(flat=True)
    return form_data or None


def _record_summary(record: OccurrenceRecord) -> str:
    lines = [
        f"Registro ID: {record.id}",
        f"Matrícula: {record.matricula or '-'}",
        f"Colaborador: {record.nome}",
        f"Setor: {record.setor}",
        f"Cargo: {record.cargo}",
        f"Turno: {record.turno}",
        f"Supervisor: {record.supervisor}",
        f"Motivo: {record.motivo or '-'}",
        f"Grau: {record.grau_label or '-'}",
        f"Grau (valor): {record.grau if record.grau is not None else '-'}",
        f"Volumes: {record.volumes if record.volumes is not None else '-'}",
        f"Observação: {record.observacao or '-'}",
        f"Criado em: {record.created_at.isoformat()}",
    ]
    return "\n".join(lines)


def _serialize_occurrence_record(record: OccurrenceRecord) -> Dict[str, Any]:
    return {
        "id": record.id,
        "matricula": record.matricula,
        "nome": record.nome,
        "setor": record.setor,
        "cargo": record.cargo,
        "turno": record.turno,
        "supervisor": record.supervisor,
        "motivo": record.motivo,
        "grau": record.grau,
        "grau_label": record.grau_label,
        "volumes": record.volumes,
        "observacao": record.observacao,
        "created_at": record.created_at.isoformat(),
    }
